from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import ObstacleBox, Waypoint, WorkbookSettings
from .validation import DataValidationError, validate_and_build_problem


TRUE_VALUES = {"是", "yes", "true", "1", "启用", "y"}


def _normalize_distance_mode(value: object) -> str:
    text = str(value or "欧氏距离").strip().lower()
    mapping = {
        "欧氏距离": "euclidean",
        "euclidean": "euclidean",
        "平面距离": "euclidean",
        "经纬度距离": "haversine",
        "haversine": "haversine",
        "球面距离": "haversine",
    }
    if text not in mapping:
        raise DataValidationError(f"无法识别距离模式：{value}")
    return mapping[text]


def _optional_float(value: object) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise DataValidationError(f"无法把“{value}”识别为数字。") from exc
    if not math.isfinite(result):
        raise DataValidationError(f"数值“{value}”不是有限数字。")
    return result


def _optional_int(value: object) -> int | None:
    number = _optional_float(value)
    if number is None:
        return None
    if not float(number).is_integer():
        raise DataValidationError(f"数值“{value}”必须是整数。")
    return int(number)


def _load_chinese_template(
    workbook: Any,
) -> tuple[list[Waypoint], WorkbookSettings, list[ObstacleBox]]:
    waypoint_sheet = workbook["航点数据"]
    headers = {
        str(value).strip(): index
        for index, value in enumerate(
            next(
                waypoint_sheet.iter_rows(
                    min_row=3,
                    max_row=3,
                    values_only=True,
                )
            )
        )
        if value is not None
    }

    def column(row: tuple[object, ...], *names: str, default=None):
        for name in names:
            if name in headers and headers[name] < len(row):
                return row[headers[name]]
        return default

    points: list[Waypoint] = []
    for row_number, row in enumerate(
        waypoint_sheet.iter_rows(min_row=4, values_only=True),
        start=4,
    ):
        waypoint_id = column(row, "编号")
        name = column(row, "名称")
        x = column(row, "X或纬度")
        y = column(row, "Y或经度")
        z = column(row, "Z或高度", default=0)
        demand = column(row, "需求量", default=0)
        enabled = column(row, "是否启用", default="是")
        note = column(row, "备注", default="")
        if waypoint_id is None and x is None and y is None:
            continue
        if str(enabled or "是").strip().lower() not in TRUE_VALUES:
            continue
        try:
            numeric_id = float(waypoint_id)
        except (TypeError, ValueError) as exc:
            raise DataValidationError(f"航点数据第 {row_number} 行的编号必须是整数。") from exc
        if not numeric_id.is_integer():
            raise DataValidationError(f"航点数据第 {row_number} 行的编号必须是整数。")
        points.append(
            Waypoint(
                waypoint_id=int(numeric_id),
                name=str(name or f"航点{int(numeric_id)}").strip(),
                x=x,
                y=y,
                z=0 if z is None else z,
                demand=0 if demand is None else demand,
                note=str(note or "").strip(),
            )
        )

    parameter_sheet = workbook["任务参数"]
    task_values: dict[str, object] = {}
    for row in parameter_sheet.iter_rows(min_row=4, max_col=2, values_only=True):
        key, value = row
        if key is not None:
            task_values[str(key).strip()] = value

    problem_type = str(task_values.get("问题类型") or "CDVRP").strip().upper()
    algorithm = str(task_values.get("算法") or "ACO").strip().upper()
    seed = _optional_int(task_values.get("随机种子"))
    settings = WorkbookSettings(
        problem_type=problem_type,
        algorithm=algorithm,
        distance_mode=_normalize_distance_mode(task_values.get("距离模式")),
        distance_unit=str(task_values.get("距离单位") or "km").strip(),
        dimension=str(task_values.get("空间维度") or "2D").strip().upper(),
        capacity=_optional_float(task_values.get("单机容量")),
        max_route_distance=_optional_float(task_values.get("单机最大航程")),
        max_vehicles=_optional_int(task_values.get("最大无人机数量")),
        seed=42 if seed is None else seed,
        min_flight_altitude=_optional_float(task_values.get("最小飞行高度")),
        max_flight_altitude=_optional_float(task_values.get("最大飞行高度")),
        obstacle_clearance=(
            _optional_float(task_values.get("障碍物安全距离")) or 0.0
        ),
    )

    algorithm_parameters: dict[str, dict[str, float]] = {}
    if "算法参数" in workbook.sheetnames:
        algorithm_sheet = workbook["算法参数"]
        for row in algorithm_sheet.iter_rows(min_row=4, max_col=4, values_only=True):
            algorithm_name, key, _label, value = row
            if algorithm_name is None or key is None or value is None:
                continue
            try:
                numeric_value = float(value)
            except (TypeError, ValueError) as exc:
                raise DataValidationError(
                    f"算法 {algorithm_name} 的参数 {key} 必须是数字。"
                ) from exc
            algorithm_parameters.setdefault(
                str(algorithm_name).strip().upper(),
                {},
            )[str(key).strip()] = numeric_value
    settings.algorithm_parameters = algorithm_parameters

    obstacles: list[ObstacleBox] = []
    if "障碍物区域" in workbook.sheetnames:
        obstacle_sheet = workbook["障碍物区域"]
        for row_number, row in enumerate(
            obstacle_sheet.iter_rows(min_row=4, max_col=10, values_only=True),
            start=4,
        ):
            (
                obstacle_id,
                obstacle_name,
                x_min,
                x_max,
                y_min,
                y_max,
                z_min,
                z_max,
                enabled,
                note,
            ) = row
            if obstacle_id is None and x_min is None and y_min is None:
                continue
            if str(enabled or "是").strip().lower() not in TRUE_VALUES:
                continue
            try:
                numeric_id = float(obstacle_id)
            except (TypeError, ValueError) as exc:
                raise DataValidationError(
                    f"障碍物区域第 {row_number} 行的编号必须是整数。"
                ) from exc
            if not numeric_id.is_integer():
                raise DataValidationError(
                    f"障碍物区域第 {row_number} 行的编号必须是整数。"
                )
            obstacles.append(
                ObstacleBox(
                    obstacle_id=int(numeric_id),
                    name=str(
                        obstacle_name or f"障碍物{int(numeric_id)}"
                    ).strip(),
                    x_min=x_min,
                    x_max=x_max,
                    y_min=y_min,
                    y_max=y_max,
                    z_min=z_min,
                    z_max=z_max,
                    note=str(note or "").strip(),
                )
            )
    return points, settings, obstacles


def _sheet_column_values(sheet: Any, column_index: int = 0) -> list[object]:
    values: list[object] = []
    for row in sheet.iter_rows(values_only=True):
        if len(row) <= column_index:
            continue
        value = row[column_index]
        if value is not None:
            values.append(value)
    return values


def _load_legacy_template(
    workbook: Any,
) -> tuple[list[Waypoint], WorkbookSettings, list[ObstacleBox]]:
    city_rows = [
        row[:2]
        for row in workbook["City"].iter_rows(values_only=True)
        if len(row) >= 2 and row[0] is not None and row[1] is not None
    ]
    if not city_rows:
        raise DataValidationError("旧格式 City 工作表中没有有效坐标。")

    demand_values = (
        _sheet_column_values(workbook["Demand"])
        if "Demand" in workbook.sheetnames
        else [0] * len(city_rows)
    )
    if len(demand_values) != len(city_rows):
        raise DataValidationError("旧格式 City 与 Demand 的行数不一致。")

    points = [
        Waypoint(
            waypoint_id=index,
            name="基地" if index == 0 else f"任务点{index}",
            x=row[0],
            y=row[1],
            demand=demand_values[index],
        )
        for index, row in enumerate(city_rows)
    ]
    has_cdvrp = "Capacity" in workbook.sheetnames and "Travelcon" in workbook.sheetnames
    capacity = (
        _optional_float(_sheet_column_values(workbook["Capacity"])[0])
        if has_cdvrp
        else None
    )
    max_route_distance = (
        _optional_float(_sheet_column_values(workbook["Travelcon"])[0])
        if has_cdvrp
        else None
    )
    settings = WorkbookSettings(
        problem_type="CDVRP" if has_cdvrp else "TSP",
        algorithm="ACO",
        distance_mode="euclidean",
        distance_unit="km",
        capacity=capacity,
        max_route_distance=max_route_distance,
        max_vehicles=None,
        seed=42,
    )
    return points, settings, []


def load_planning_workbook(
    input_path: str | Path,
    overrides: dict[str, object] | None = None,
):
    """读取新中文模板或原 MATLAB 四工作表格式。"""

    path = Path(input_path).expanduser()
    if not path.exists():
        raise DataValidationError(f"找不到 Excel 文件：{path.name}")
    if path.suffix.lower() != ".xlsx":
        raise DataValidationError("输入文件必须是 .xlsx 格式。")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except PermissionError as exc:
        raise DataValidationError("无法读取 Excel；请先保存并关闭 Excel/WPS 后重试。") from exc
    except Exception as exc:
        raise DataValidationError(f"Excel 文件无法打开：{exc}") from exc

    try:
        if {"航点数据", "任务参数"}.issubset(workbook.sheetnames):
            points, settings, obstacles = _load_chinese_template(workbook)
        elif "City" in workbook.sheetnames:
            points, settings, obstacles = _load_legacy_template(workbook)
        else:
            raise DataValidationError(
                "工作簿格式不受支持：需要“航点数据/任务参数”或“City”工作表。"
            )
    finally:
        workbook.close()

    for key, value in (overrides or {}).items():
        if value is not None and hasattr(settings, key):
            setattr(settings, key, value)

    if settings.algorithm not in {"ACO", "GA", "HPSO", "SA"}:
        raise DataValidationError("算法只能是 ACO、GA、HPSO 或 SA。")
    if settings.seed < 0:
        raise DataValidationError("随机种子不能为负数。")

    problem = validate_and_build_problem(points, settings, obstacles)
    return problem, settings
